import pandas as pd
from openpyxl.utils import get_column_letter
from pulp import *
from openpyxl import Workbook

dir_data_prod="data/prod"
dir_data_test="data/test"
dir_data_in_use = dir_data_test

file_src = 'DEA.txt'
file_stg = 'DEA_stg.txt'
file_in_use = file_src

file_preprocessed = 'DEA_processed.csv'
file_optimized_weight_factor = 'DEA_weightFactors.csv'
file_optimized_LPModel = 'DEA_LPModel.csv'

file_dea_report = 'DEA_Report.xlsx'

df_input = None
df_output = None

# Load data
def load_data():
    lines = None
    lines_stg = []
    global file_in_use
    with open(f'{dir_data_in_use}/{file_in_use}', 'r') as file:
        lines = file.readlines()

    for line in lines:
        if line == '\n':
            pass
        else:
            lines_stg.append(line.strip())
    print(lines_stg)

    file_in_use = file_stg
    with open(f'{dir_data_in_use}/{file_in_use}', 'w') as file:
        for line in lines_stg:
            file.write(f"{line}\n")
    print('**********************loading data complete**************************')


# pre-process data
def preprocess_data():
    print('preprocessing data')
    global file_in_use
    global inputName
    global OutputName
    file_in_use = file_stg
    unitName = []
    inputUsed = []
    OutputProduced = []
    with open(f'{dir_data_in_use}/{file_in_use}', 'r') as f:
        line = f.readline().strip()
        inputName = line.split(',')
        print(inputName)
        line = f.readline().strip()
        OutputName = line.split(',')
        print(OutputName)
        while f :
            line = f.readline().strip()
            if line == '':
                break
            unitName.append(line)
            line = f.readline().strip()
            line = line.split(',')
            line = [float(i) for i in line]
            inputUsed.append(line)
            line = f.readline().strip()
            line = line.split(',')
            line = [float(i) for i in line]
            OutputProduced.append(line)
    print(unitName)
    print(inputUsed)
    print(OutputProduced)

    df_input = pd.DataFrame(data = inputUsed, columns=inputName, index = unitName)
    df_output = pd.DataFrame(data = OutputProduced, columns=OutputName, index = unitName)
    print(df_input)
    print(df_output)

    pd.concat([df_input,df_output],axis=1).to_csv(f'{dir_data_in_use}/{file_preprocessed}')
    print(pd.concat([df_input, df_output],axis=1))

    print("***********************preprocess_data complete**********************")

    # build_model():  &  execute_model():
def build_model():
    global file_in_use
    file_in_use = file_preprocessed
    df = pd.read_csv(f'{dir_data_in_use}/{file_in_use}',index_col = 0)
    print(df)
    print('build_model')
    print(inputName)

    in_vars = LpVariable.dicts('I',inputName, lowBound=0, cat='Continuous')
    out_vars = LpVariable.dicts('O',OutputName, lowBound=0, cat='Continuous')
    df_weightFactor = pd.DataFrame(index = df.index, columns = [inputName + OutputName])
    print(df_weightFactor)

    print(in_vars, out_vars)
    print('**************************************')
    for unit in df.index:
        prob = LpProblem("DEA Analysis", LpMaximize)

        for j in range(df.shape[0]):
            prob += (lpSum([df.loc[df.index[j],i] * in_vars[i] for i in inputName]) >=
                    lpSum([df.loc[df.index[j],i] * out_vars[i] for i in OutputName])
                    )

        prob += lpSum([df.loc[unit, i] * in_vars[i] for i in inputName]) == 1
        prob += lpSum([df.loc[unit, i] * out_vars[i] for i in OutputName])

        print(df.index)
        status = prob.solve()

        weightFactorDict = {}
        for v in prob.variables():
            weightFactorDict[v.name] = v.varValue
            print(v.name, v.varValue)

        df_weightFactor.loc[unit]=[weightFactorDict['I_Faculty'],
                                 weightFactorDict['I_Support_Staff'],
                                 weightFactorDict['I_Supply_Budget'],
                                 weightFactorDict['O_Credit_Hours'],
                                 weightFactorDict['O_Research_Pubs']
                                 ]

    print(df_weightFactor)
    df_weightFactor.to_csv(f'{dir_data_in_use}/{file_optimized_weight_factor}')

# generate report
def prepare_report_data():
    print('****************prepare_report_data*************************')
    global file_in_use
    file_in_use = file_optimized_weight_factor
    df_weightFactor = pd.read_csv(f'{dir_data_in_use}/{file_in_use}',index_col = 0)

    file_in_use = file_preprocessed
    df_preprocessed = pd.read_csv(f'{dir_data_in_use}/{file_in_use}',index_col = 0)
    print(df_preprocessed)

    df_LPModel = df_weightFactor * df_preprocessed
    df_LPModel.to_csv(f'{dir_data_in_use}/{file_optimized_LPModel}')
    print(df_LPModel)
    print('****************prepare_report_data complete*************************')

def generate_report():
    wb = Workbook()
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    ws.title = 'DEA Report'

    row_num = 4
    col_num = 2
    loc_title_row = row_num -3
    loc_title_col_char = get_column_letter(col_num)
    ws[f'{loc_title_col_char}{loc_title_row}'] = 'Summary of Analysis'

    build_report_section1(ws, [row_num, col_num])
    build_report_section2(ws, [row_num, col_num+4])
    build_report_section3(ws, [row_num+7, col_num+4])
    build_report_section4(ws, [row_num+14, col_num+4])

    dir_data_in_use = dir_data_prod
    wb.save(f'{dir_data_in_use}/{file_dea_report}')

def build_report_section1(ws, loc):
    file_in_use = file_optimized_LPModel
    df_LPModel = pd.read_csv(f'{dir_data_in_use}/{file_in_use}',index_col = 0)
    df_LPModel['LPMax'] = df_LPModel[OutputName].sum(axis=1)
    df_LPModel['eff_index'] = df_LPModel['LPMax'].apply(lambda x: 'Yes' if x>=1 else 'No')
    df_eff = df_LPModel[['LPMax','eff_index']]
    print(df_eff)
    row_num, col_num = loc

    loc_eff_row = row_num
    loc_eff_col_char = get_column_letter(col_num)

    ws[f'{loc_eff_col_char}{loc_eff_row}'] = 'Units'
    loc_eff_col_char = get_column_letter(col_num+1)
    ws[f'{loc_eff_col_char}{loc_eff_row}'] = 'LP Maximum Output'
    loc_eff_col_char = get_column_letter(col_num + 2)
    ws[f'{loc_eff_col_char}{loc_eff_row}'] = 'Efficient?'

    for i in range(df_eff.shape[0]):
        loc_eff_col_char = get_column_letter(col_num)
        ws[f'{loc_eff_col_char}{loc_eff_row + i + 1}'] = df_eff.index[i]

        loc_eff_col_char = get_column_letter(col_num+1)
        ws[f'{loc_eff_col_char}{loc_eff_row + i + 1}'] = df_eff.loc[df_eff.index[i], 'LPMax']

        loc_eff_col_char = get_column_letter(col_num+2)
        ws[f'{loc_eff_col_char}{loc_eff_row + i + 1}'] = df_eff.loc[df_eff.index[i], 'eff_index']

def build_report_section2(ws, loc):
    file_in_use = file_preprocessed
    df = pd.read_csv(f'{dir_data_in_use}/{file_in_use}', index_col=0)
    row_num, col_num = loc

    loc_row = row_num
    loc_col_char = get_column_letter(col_num)

    ws[f'{loc_col_char}{loc_row}'] = 'Units'

    for i in range(len(inputName)):
        loc_col_char = get_column_letter(col_num+i+1)
        ws[f'{loc_col_char}{loc_row}'] = inputName[i]
        for j in range(df.shape[0]):
            ws[f'{loc_col_char}{loc_row+j+1}'] = df.loc[df.index[j], inputName[i]]

    for i in range(len(OutputName)):
        loc_col_char = get_column_letter(col_num+i+1+len(inputName))
        ws[f'{loc_col_char}{loc_row}'] = OutputName[i]
        for j in range(df.shape[0]):
            ws[f'{loc_col_char}{loc_row+j+1}'] = df.loc[df.index[j], OutputName[i]]

    for i in range(df.shape[0]):
        loc_col_char = get_column_letter(col_num)
        ws[f'{loc_col_char}{loc_row+i+1}'] = df.index[i]

def build_report_section3(ws, loc):
    file_in_use = file_optimized_weight_factor
    df = pd.read_csv(f'{dir_data_in_use}/{file_in_use}', index_col=0)
    row_num, col_num = loc

    loc_row = row_num
    loc_col_char = get_column_letter(col_num)

    ws[f'{loc_col_char}{loc_row}'] = 'Units'

    for i in range(len(inputName)):
        loc_col_char = get_column_letter(col_num+i+1)
        ws[f'{loc_col_char}{loc_row}'] = inputName[i]
        for j in range(df.shape[0]):
            ws[f'{loc_col_char}{loc_row+j+1}'] = df.loc[df.index[j], inputName[i]]

    for i in range(len(OutputName)):
        loc_col_char = get_column_letter(col_num+i+1+len(inputName))
        ws[f'{loc_col_char}{loc_row}'] = OutputName[i]
        for j in range(df.shape[0]):
            ws[f'{loc_col_char}{loc_row+j+1}'] = df.loc[df.index[j], OutputName[i]]

    for i in range(df.shape[0]):
        loc_col_char = get_column_letter(col_num)
        ws[f'{loc_col_char}{loc_row+i+1}'] = df.index[i]

def build_report_section4(ws, loc):
    file_in_use = file_optimized_LPModel
    df = pd.read_csv(f'{dir_data_in_use}/{file_in_use}', index_col=0)
    row_num, col_num = loc

    loc_row = row_num
    loc_col_char = get_column_letter(col_num)

    ws[f'{loc_col_char}{loc_row}'] = 'Units'

    for i in range(len(inputName)):
        loc_col_char = get_column_letter(col_num + i + 1)
        ws[f'{loc_col_char}{loc_row}'] = inputName[i]
        for j in range(df.shape[0]):
            ws[f'{loc_col_char}{loc_row + j + 1}'] = df.loc[df.index[j], inputName[i]]

    for i in range(len(OutputName)):
        loc_col_char = get_column_letter(col_num + i + 1 + len(inputName))
        ws[f'{loc_col_char}{loc_row}'] = OutputName[i]
        for j in range(df.shape[0]):
            ws[f'{loc_col_char}{loc_row + j + 1}'] = df.loc[df.index[j], OutputName[i]]

    for i in range(df.shape[0]):
        loc_col_char = get_column_letter(col_num)
        ws[f'{loc_col_char}{loc_row + i + 1}'] = df.index[i]


if __name__ == '__main__':
    load_data()
    preprocess_data()
    build_model()
    prepare_report_data()
    generate_report()